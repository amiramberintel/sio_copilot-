#!/nfs/site/disks/ayarokh_wa/tools/python/3.11.1/bin/python3

import os
import sys
import pwd
import json
import openpyxl
import argparse
import xlsxwriter
from io import BytesIO
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.client_credential import ClientCredential


version = 'V0.1a'


class sio_uploader:
    ctx = None
    mode = None
    config = None
    config_name = 'sio_uploader.json'

    def __init__(self, mode):
        self.mode = mode
        self.config = self.load_config()
        self.ctx = ClientContext(self.config['urls'][self.mode]['sp_site']).with_credentials(
            ClientCredential(self.config['data']['client_id'], self.config['data']['secret']))

    def load_config(self):
        d = os.path.dirname(os.path.realpath(__file__))
        f = os.path.join(d, self.config_name)
        config = None
        with open(f, 'r') as openfile:
            config = json.load(openfile)
        return config

    def create_excel_fadi_row(self, workbook, infile, fields, worksheet_name, add_to_table=[]):
        worksheet = workbook.add_worksheet(worksheet_name)
        with open(infile) as f:
            row = 0
            for line in f:
                if line.strip() == "":
                    continue
                row += 1
                start = 0
                lline = []
                for i in range(0, len(fields)-1):
                    end = line.find('|', start)
                    d = line[start: end].strip()
                    if d != "" and i > 0:
                        try:
                            d = float(d)
                        except ValueError:
                            pass
                    lline.append(d)
                    start = end + 1
                lline.append(line[start:].strip())
                worksheet.write_row(row, 0, lline)
            columns = [{'header': l} for l in fields]
            columns.extend(add_to_table)
            worksheet.add_table(0, 0, row, len(
                fields)+len(add_to_table)-1, {"name": f"tbl_{worksheet_name}", 'columns': columns})
            return worksheet

    def create_excel_fadi_normalized(self, workbook, infile):
        fields = ['port', 'TNS', 'new_WNS', 'ref_WNS', 'comment']
        vl = f'VLOOKUP([@port],tbl_mow[],4,FALSE)'
        add = [{'header': 'norm_comment', 'formula': f'=IF({vl}=0,"",{vl})'}]
        add.append(
            {'header': 'need to review', 'formula': f'=IF(SUBSTITUTE([@{fields[-1]}]," ","")=SUBSTITUTE([@[{add[0]["header"]}]]," ",""),"","need")'})
        self.create_excel_fadi_row(
            workbook, infile, fields, "mow_normilized", add)

    def create_excel_fadi(self, workbook, infile):
        fields = ['port', 'new_WNS', 'ref_WNS', 'comment']

        vl = f'VLOOKUP([@port],tbl_mow_normilized[],5,FALSE)'
        add = [{'header': 'flat_comment', 'formula': f'=IF({vl}=0,"",{vl})'}]
        add.append(
            {'header': 'need to review', 'formula': f'=IF(SUBSTITUTE([@{fields[-1]}]," ","")=SUBSTITUTE([@[{add[0]["header"]}]]," ",""),"","need")'})
        self.create_excel_fadi_row(workbook, infile, fields, "mow", add)

    def upload(self, data, file_name):
        target_folder = self.ctx.web.get_folder_by_server_relative_url(
            self.config['urls'][self.mode]['folder'])
        uploaded_file = target_folder.upload_file(
            file_name, data).execute_query()
        return uploaded_file.serverRelativeUrl

    def download(self, file_name, output_dir):
        boutput = BytesIO()
        basename = os.path.splitext(file_name)[0]
        file = self.ctx.web.get_file_by_server_relative_url(
            os.path.join(self.config['urls'][self.mode]["folder"], file_name)).download(boutput).execute_query()
        wb = openpyxl.load_workbook(boutput, read_only=True)
        for sheetname in wb.sheetnames:
            output = []
            sheet = wb[sheetname]
            for row in range(2, sheet.max_row+1):
                l = []
                for col in range(1, sheet.max_column-1):
                    v = sheet.cell(column=col, row=row).value
                    if v is None:
                        v = ""
                    l.append(str(v))
                output.append(" | ".join(l))
            fout = f'{output_dir}/{basename}_{sheetname}.txt'
            with open(fout, mode='w') as out:
                out.write("\n".join(output))
            print(f'File wrote: {fout}')
        boutput.close()


def parse_args():
    """ parse_args() : command-line parser """
    parser = argparse.ArgumentParser(
        description="Get file and upload it to user's sharepoint")
    parser.add_argument('mode',  option_strings=[
                        'demo', 'fadi'])
    group = parser.add_mutually_exclusive_group(required=True)
    parser.add_argument('-output_dir', type=str,
                        help="Output directory for '-get_file' option", default="./", required=False)
    group.add_argument('-mow_file', type=str, help="upload MOW file")
    group.add_argument('-get_file', type=str, help="get MOW file")
    args = parser.parse_args()
    return args


def main():
    args = parse_args()
    current_user = pwd.getpwuid(os.getuid())[0]
    if args.mode in ['fadi', 'demo']:
        sio_up = sio_uploader(args.mode)
        if args.mow_file is not None:
            output = BytesIO()
            workbook = xlsxwriter.Workbook(output)
            tt = os.path.splitext(args.mow_file)
            sio_up.create_excel_fadi(workbook, args.mow_file)
            sio_up.create_excel_fadi_normalized(
                workbook, f'{tt[0]}_normalized{tt[1]}')
            workbook.set_custom_property('Checked by', current_user)
            workbook.set_custom_property('Version', version)
            workbook.set_custom_property(
                'Original_path', os.path.realpath(args.mow_file))
            workbook.close()
            ret = sio_up.upload(
                output.getvalue(), f'{os.path.basename(tt[0])}.xlsx')
            if ret is not None:
                print(f'File uploded: {ret}')
                return 0
            else:
                print(f'File failes to upload')
                return 1
        if args.get_file is not None:
            sio_up.download(args.get_file, args.output_dir)
    else:
        raise Exception(f'Unknown mode: {args.mode}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
